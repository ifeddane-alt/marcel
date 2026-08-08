"""Export/Import Excel générique — .xlsx lisible directement, import avec upsert."""
import io
import re
import uuid
from datetime import datetime, timezone, date as date_cls

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.database import db
from core.auth import TokenPayload, require_write
from .specs import SPECS

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _norm(s) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(s or "").casefold())


def _key(s) -> str:
    return str(s or "").strip().casefold()


def _spec(entity: str) -> dict:
    if entity not in SPECS:
        raise HTTPException(status_code=404, detail=f"Entité inconnue : {entity}")
    return SPECS[entity]


# ─── Parsing des valeurs importées ──────────────────────────────────────────

def _parse_value(raw, c):
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        return None, None
    t = c["type"]
    if t == "date":
        if isinstance(raw, datetime):
            return raw.strftime("%Y-%m-%d"), None
        if isinstance(raw, date_cls):
            return raw.isoformat(), None
        s = str(raw).strip()[:10]
        try:
            return date_cls.fromisoformat(s).isoformat(), None
        except ValueError:
            pass
        try:
            return datetime.strptime(str(raw).strip(), "%d/%m/%Y").strftime("%Y-%m-%d"), None
        except ValueError:
            return None, f"{c['label']} : date invalide « {raw} » (attendu JJ/MM/AAAA)"
    if t in ("money", "number", "int"):
        s = str(raw).replace("€", "").replace("%", "").replace("\u202f", "").replace("\xa0", "")
        s = s.replace(" ", "").replace(",", ".")
        try:
            v = float(s)
        except ValueError:
            return None, f"{c['label']} : valeur non numérique « {raw} »"
        return (int(v) if t == "int" else v), None
    if t == "bool":
        s = str(raw).strip().casefold()
        if s in ("oui", "yes", "true", "vrai", "1", "x"):
            return True, None
        if s in ("non", "no", "false", "faux", "0", ""):
            return False, None
        return None, f"{c['label']} : valeur booléenne invalide « {raw} » (attendu Oui/Non)"
    if t == "enum":
        s = _key(raw)
        for canon, label in c["options"].items():
            if s in (canon.casefold(), label.casefold()):
                return canon, None
        for alias, canon in (c.get("aliases") or {}).items():
            if s == alias.casefold():
                return canon, None
        allowed = ", ".join(c["options"].values())
        return None, f"{c['label']} : valeur invalide « {raw} » (attendu : {allowed})"
    return str(raw).strip(), None


def _display(v, c):
    if v is None:
        return None
    t = c["type"]
    if t == "enum":
        return c["options"].get(v, v)
    if t == "bool":
        return "Oui" if v else "Non"
    if t == "date":
        try:
            return datetime.strptime(str(v)[:10], "%Y-%m-%d")
        except ValueError:
            return str(v)
    return v


# ─── Export ──────────────────────────────────────────────────────────────────

THIN = Border(*[Side(style="thin", color="D9D9D9")] * 4)


def _build_workbook(spec: dict, rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = spec["label"][:31]
    cols = spec["columns"]

    hdr_fill = PatternFill("solid", fgColor="2563EB")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill = PatternFill("solid", fgColor="F4F7FB")

    for ci, c in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=ci, value=c["label"])
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN

    widths = [len(c["label"]) + 4 for c in cols]
    for ri, row in enumerate(rows, start=2):
        for ci, c in enumerate(cols, start=1):
            v = _display(row.get(c["field"]), c)
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = THIN
            if ri % 2 == 0:
                cell.fill = alt_fill
            if c["type"] == "money":
                cell.number_format = '#,##0" €"'
            elif c["type"] == "number":
                cell.number_format = "#,##0.##"
            elif c["type"] == "date":
                cell.number_format = "DD/MM/YYYY"
            widths[ci - 1] = min(max(widths[ci - 1], len(str(v or "")) + 3), 55)

    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{max(len(rows) + 1, 2)}"

    # Onglet d'aide pour le ré-import
    help_ws = wb.create_sheet("Aide import")
    help_ws.append(["Colonne", "Obligatoire (création)", "Format", "Valeurs autorisées"])
    for ci in range(1, 5):
        hc = help_ws.cell(row=1, column=ci)
        hc.fill = hdr_fill
        hc.font = hdr_font
    fmt_hint = {"date": "JJ/MM/AAAA", "money": "Nombre (€)", "number": "Nombre",
                "int": "Entier", "bool": "Oui / Non", "enum": "Liste", "str": "Texte"}
    for c in spec["columns"]:
        help_ws.append([
            c["label"] + (" (lecture seule)" if c.get("readonly") else ""),
            "Oui" if c["field"] in spec.get("required_new", []) else "Non",
            fmt_hint.get(c["type"], "Texte"),
            ", ".join(c["options"].values()) if c["type"] == "enum" else "",
        ])
    help_ws.column_dimensions["A"].width = 30
    help_ws.column_dimensions["B"].width = 22
    help_ws.column_dimensions["C"].width = 16
    help_ws.column_dimensions["D"].width = 60

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


async def _tenant_projects(user) -> list[dict]:
    return await db.projects.find({"tenant_id": user.tenant_id}, {"_id": 0}).to_list(None)


async def _fetch_rows(entity: str, user: TokenPayload) -> list[dict]:
    tid = user.tenant_id
    if entity == "projects" or entity == "budget":
        projects = await _tenant_projects(user)
        programs = await db.programs.find({"tenant_id": tid}, {"_id": 0}).to_list(None)
        prog_map = {p["program_id"]: p["name"] for p in programs}
        rows = []
        for p in sorted(projects, key=lambda x: _key(x.get("name"))):
            if entity == "projects":
                rows.append({**p, "program_name": prog_map.get(p.get("program_id"), "")})
            else:
                cp, cc = p.get("capex_planned") or 0, p.get("capex_consumed") or 0
                op, oc = p.get("opex_planned") or 0, p.get("opex_consumed") or 0
                eac = p.get("eac") or p.get("budget_forecast") or (cp + op)
                budget = cp + op
                rows.append({
                    "name": p["name"], "program_name": prog_map.get(p.get("program_id"), ""),
                    "capex_planned": cp, "capex_consumed": cc,
                    "opex_planned": op, "opex_consumed": oc, "eac": eac,
                    "raf": max(eac - cc - oc, 0),
                    "ecart_pct": round((eac - budget) / budget * 100, 1) if budget else 0,
                })
        return rows
    if entity == "programs":
        progs = await db.programs.find({"tenant_id": tid}, {"_id": 0}).to_list(None)
        return sorted(progs, key=lambda x: _key(x.get("name")))
    if entity == "teams":
        teams = await db.teams.find({"tenant_id": tid}, {"_id": 0}).to_list(None)
        resources = await db.resources.find({"tenant_id": tid}, {"_id": 0}).to_list(None)
        res_map = {r["resource_id"]: r["name"] for r in resources}
        counts: dict = {}
        for r in resources:
            if r.get("team_id"):
                counts[r["team_id"]] = counts.get(r["team_id"], 0) + 1
        return [{
            **t,
            "manager_name": res_map.get(t.get("manager_resource_id"), ""),
            "members_count": counts.get(t["team_id"], 0),
        } for t in sorted(teams, key=lambda x: _key(x.get("name")))]
    if entity == "resources":
        res = await db.resources.find({"tenant_id": tid}, {"_id": 0}).to_list(None)
        return sorted(res, key=lambda x: _key(x.get("name")))
    if entity == "milestones":
        projects = await _tenant_projects(user)
        pmap = {p["project_id"]: p["name"] for p in projects}
        ms = await db.milestones.find(
            {"project_id": {"$in": list(pmap.keys())}}, {"_id": 0}
        ).to_list(None)
        rows = [{**m, "project_name": pmap.get(m.get("project_id"), "")} for m in ms]
        return sorted(rows, key=lambda x: (_key(x["project_name"]), str(x.get("date_forecast") or "")))
    if entity in ("risks", "decisions"):
        coll = db.risks if entity == "risks" else db.decisions
        docs = await coll.find({"tenant_id": tid}, {"_id": 0}).to_list(None)
        projects = await _tenant_projects(user)
        pmap = {p["project_id"]: p["name"] for p in projects}
        rows = [{**d, "project_name": pmap.get(d.get("project_id"), "")} for d in docs]
        return sorted(rows, key=lambda x: (_key(x["project_name"]), _key(x.get("title"))))
    if entity == "demands":
        docs = await db.demands.find({"tenant_id": tid}, {"_id": 0}).to_list(None)
        return sorted(docs, key=lambda x: _key(x.get("title")))
    if entity == "timesheets":
        ts = await db.timesheets.find({"tenant_id": tid}, {"_id": 0}).to_list(None)
        resources = await db.resources.find({"tenant_id": tid}, {"_id": 0}).to_list(None)
        res_map = {r["resource_id"]: r["name"] for r in resources}
        wa_ids = list({t["work_allocation_id"] for t in ts})
        was = await db.work_allocations.find(
            {"work_allocation_id": {"$in": wa_ids}}, {"_id": 0}
        ).to_list(None)
        wa_map = {w["work_allocation_id"]: w for w in was}
        task_ids = list({w.get("task_id") for w in was if w.get("task_id")})
        tasks = await db.tasks.find({"task_id": {"$in": task_ids}}, {"_id": 0}).to_list(None)
        task_map = {t["task_id"]: t for t in tasks}
        projects = await _tenant_projects(user)
        pmap = {p["project_id"]: p["name"] for p in projects}
        rows = []
        for t in ts:
            wa = wa_map.get(t["work_allocation_id"], {})
            task = task_map.get(wa.get("task_id", ""), {})
            rows.append({
                "resource_name": res_map.get(t["resource_id"], ""),
                "project_name": pmap.get(task.get("project_id"), ""),
                "task_name": task.get("name", ""),
                "date": t.get("date"),
                "jh_value": t.get("jh_value", 0),
                "status": t.get("status", "draft"),
            })
        return sorted(rows, key=lambda x: (_key(x["resource_name"]), str(x["date"] or "")))
    raise HTTPException(status_code=404, detail=f"Entité inconnue : {entity}")


async def export_entity(entity: str, user: TokenPayload) -> tuple[str, bytes]:
    spec = _spec(entity)
    rows = await _fetch_rows(entity, user)
    data = _build_workbook(spec, rows)
    fname = f"MARCEL_{spec['label']}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return fname, data


# ─── Analyse (préview) et application (commit) ───────────────────────────────

async def _build_ctx(entity: str, user: TokenPayload) -> dict:
    tid = user.tenant_id
    ctx: dict = {"tenant_id": tid}
    if entity in ("projects", "budget", "milestones", "risks", "decisions", "timesheets"):
        projects = await _tenant_projects(user)
        ctx["projects_by_name"] = {_key(p["name"]): p for p in projects}
        ctx["projects_by_id"] = {p["project_id"]: p for p in projects}
    if entity == "projects":
        programs = await db.programs.find({"tenant_id": tid}, {"_id": 0}).to_list(None)
        ctx["programs_by_name"] = {_key(p["name"]): p for p in programs}
    if entity == "programs":
        programs = await db.programs.find({"tenant_id": tid}, {"_id": 0}).to_list(None)
        ctx["existing"] = {_key(p["name"]): p for p in programs}
    if entity == "teams":
        teams = await db.teams.find({"tenant_id": tid}, {"_id": 0}).to_list(None)
        ctx["existing"] = {_key(t["name"]): t for t in teams}
        resources = await db.resources.find({"tenant_id": tid}, {"_id": 0}).to_list(None)
        ctx["resources_by_name"] = {_key(r["name"]): r for r in resources}
    if entity == "resources":
        resources = await db.resources.find({"tenant_id": tid}, {"_id": 0}).to_list(None)
        ctx["existing"] = {_key(r["name"]): r for r in resources}
        teams = await db.teams.find({"tenant_id": tid}, {"_id": 0}).to_list(None)
        ctx["teams_by_name"] = {_key(t["name"]): t for t in teams}
    if entity == "milestones":
        pids = list(ctx["projects_by_id"].keys())
        ms = await db.milestones.find({"project_id": {"$in": pids}}, {"_id": 0}).to_list(None)
        ctx["existing"] = {(m["project_id"], _key(m.get("name"))): m for m in ms}
    if entity in ("risks", "decisions"):
        coll = db.risks if entity == "risks" else db.decisions
        docs = await coll.find({"tenant_id": tid}, {"_id": 0}).to_list(None)
        ctx["existing"] = {(d.get("project_id"), _key(d.get("title"))): d for d in docs}
    if entity == "demands":
        docs = await db.demands.find({"tenant_id": tid}, {"_id": 0}).to_list(None)
        ctx["existing"] = {_key(d.get("title")): d for d in docs}
    if entity == "timesheets":
        resources = await db.resources.find({"tenant_id": tid}, {"_id": 0}).to_list(None)
        ctx["resources_by_name"] = {_key(r["name"]): r for r in resources}
        was = await db.work_allocations.find({"tenant_id": tid}, {"_id": 0}).to_list(None)
        task_ids = list({w.get("task_id") for w in was if w.get("task_id")})
        tasks = await db.tasks.find({"task_id": {"$in": task_ids}}, {"_id": 0}).to_list(None)
        task_map = {t["task_id"]: t for t in tasks}
        # (resource_id, nom tâche normalisé) -> work_allocation
        wa_by_res_task: dict = {}
        for w in was:
            task = task_map.get(w.get("task_id", ""), {})
            wa_by_res_task[(w.get("resource_id"), _key(task.get("name")))] = w
        ctx["wa_by_res_task"] = wa_by_res_task
        ts = await db.timesheets.find({"tenant_id": tid}, {"_id": 0}).to_list(None)
        ctx["existing"] = {
            (t["resource_id"], t["work_allocation_id"], t.get("date")): t for t in ts
        }
    return ctx


def _analyze_row(entity: str, spec: dict, data: dict, ctx: dict) -> dict:
    """Retourne {action, errors, resolved} pour une ligne canonique."""
    errors: list[str] = []
    resolved: dict = {}

    # Résolution projet
    if "project_name" in data and entity in ("milestones", "risks", "decisions"):
        pname = data.get("project_name")
        if pname:
            proj = ctx["projects_by_name"].get(_key(pname))
            if not proj:
                errors.append(f"Projet introuvable : « {pname} »")
            else:
                resolved["project_id"] = proj["project_id"]

    existing = None
    if entity in ("projects", "budget"):
        existing = ctx["projects_by_name"].get(_key(data.get("name")))
        if entity == "budget" and data.get("name") and not existing:
            errors.append(f"Projet introuvable : « {data.get('name')} » (l'import budget ne crée pas de projet)")
    elif entity in ("programs", "teams", "resources"):
        existing = ctx["existing"].get(_key(data.get("name")))
    elif entity == "demands":
        existing = ctx["existing"].get(_key(data.get("title")))
    elif entity in ("milestones",):
        pid = resolved.get("project_id")
        if pid:
            existing = ctx["existing"].get((pid, _key(data.get("name"))))
    elif entity in ("risks", "decisions"):
        pid = resolved.get("project_id")
        if pid:
            existing = ctx["existing"].get((pid, _key(data.get("title"))))
    elif entity == "timesheets":
        res = ctx["resources_by_name"].get(_key(data.get("resource_name")))
        if not res:
            errors.append(f"Ressource introuvable : « {data.get('resource_name')} »")
        else:
            resolved["resource_id"] = res["resource_id"]
            wa = ctx["wa_by_res_task"].get((res["resource_id"], _key(data.get("task_name"))))
            if not wa:
                errors.append(
                    f"Aucune allocation pour « {data.get('resource_name')} » sur la tâche « {data.get('task_name')} »"
                )
            else:
                resolved["work_allocation_id"] = wa["work_allocation_id"]
                existing = ctx["existing"].get(
                    (res["resource_id"], wa["work_allocation_id"], data.get("date"))
                )
                if existing and existing.get("status") in ("submitted", "cp_reviewed", "validated"):
                    errors.append("Entrée déjà soumise ou validée : non modifiable")

    action = "update" if existing else "new"
    if action == "new":
        for req in spec.get("required_new", []):
            if data.get(req) in (None, ""):
                label = next((c["label"] for c in spec["columns"] if c["field"] == req), req)
                errors.append(f"Champ requis manquant : {label}")
        if entity == "projects" and data.get("start_date") and data.get("end_date_forecast"):
            pass
    # Résolutions annexes
    if entity == "projects" and data.get("program_name"):
        prog = ctx["programs_by_name"].get(_key(data["program_name"]))
        if prog:
            resolved["program_id"] = prog["program_id"]
        else:
            errors.append(f"Programme introuvable : « {data['program_name']} »")
    if entity == "teams" and data.get("manager_name"):
        mgr = ctx["resources_by_name"].get(_key(data["manager_name"]))
        if mgr:
            resolved["manager_resource_id"] = mgr["resource_id"]
        else:
            errors.append(f"Manager introuvable dans les ressources : « {data['manager_name']} »")
    if entity == "resources" and data.get("team"):
        team = ctx["teams_by_name"].get(_key(data["team"]))
        if team:
            resolved["team_id"] = team["team_id"]

    return {
        "action": "error" if errors else action,
        "errors": errors,
        "existing": existing,
        "resolved": resolved,
    }


def _provided(data: dict, spec: dict) -> dict:
    """Champs non-readonly effectivement fournis (non None)."""
    ro = {c["field"] for c in spec["columns"] if c.get("readonly")}
    return {k: v for k, v in data.items() if v is not None and k not in ro}


async def _apply_row(entity: str, spec: dict, data: dict, info: dict, user: TokenPayload) -> str:
    """Applique une ligne (create/update). Retourne 'created' ou 'updated'."""
    now = _now()
    tid = user.tenant_id
    existing = info["existing"]
    resolved = info["resolved"]
    fields = _provided(data, spec)

    if entity in ("projects", "budget"):
        if entity == "budget":
            upd = {k: fields[k] for k in
                   ("capex_planned", "capex_consumed", "opex_planned", "opex_consumed", "eac")
                   if k in fields}
            cp = upd.get("capex_planned", existing.get("capex_planned") or 0)
            op = upd.get("opex_planned", existing.get("opex_planned") or 0)
            upd["budget_total"] = (cp or 0) + (op or 0)
            if "eac" in upd:
                upd["budget_forecast"] = upd["eac"]
            await db.projects.update_one({"project_id": existing["project_id"]}, {"$set": upd})
            return "updated"
        payload = {k: v for k, v in fields.items() if k != "program_name"}
        if "program_id" in resolved:
            payload["program_id"] = resolved["program_id"]
        if existing:
            await db.projects.update_one({"project_id": existing["project_id"]}, {"$set": payload})
            return "updated"
        doc = {
            "project_id": str(uuid.uuid4()), "tenant_id": tid,
            "methodology": "waterfall", "status_rag": "green", "status": "actif",
            "budget_total": 0, "budget_consumed": 0, "budget_forecast": 0,
            "jh_planned": 0, "jh_consumed": 0, "metadata": {}, "created_at": now,
            **payload,
        }
        doc.setdefault("end_date_baseline", doc.get("end_date_forecast"))
        if not doc.get("budget_forecast"):
            doc["budget_forecast"] = doc.get("budget_total", 0)
        from modules.projects.service import generate_project_code
        doc["code"] = await generate_project_code(tid, doc.get("program_id"))
        await db.projects.insert_one(doc)
        return "created"

    if entity == "programs":
        if existing:
            await db.programs.update_one({"program_id": existing["program_id"]}, {"$set": fields})
            return "updated"
        doc = {"program_id": str(uuid.uuid4()), "tenant_id": tid, "budget_keur": 0,
               "status": "active", "created_at": now, **fields}
        await db.programs.insert_one(doc)
        return "created"

    if entity == "teams":
        payload = {k: v for k, v in fields.items() if k not in ("manager_name", "members_count")}
        if "manager_resource_id" in resolved:
            payload["manager_resource_id"] = resolved["manager_resource_id"]
        if existing:
            if payload:
                await db.teams.update_one({"team_id": existing["team_id"]}, {"$set": payload})
            return "updated"
        doc = {"team_id": str(uuid.uuid4()), "tenant_id": tid, "created_at": now, **payload}
        await db.teams.insert_one(doc)
        return "created"

    if entity == "resources":
        payload = dict(fields)
        if "team_id" in resolved:
            payload["team_id"] = resolved["team_id"]
        if existing:
            await db.resources.update_one({"resource_id": existing["resource_id"]}, {"$set": payload})
            return "updated"
        doc = {"resource_id": str(uuid.uuid4()), "tenant_id": tid,
               "capacity_jh_month": 15, "availability_rate": 100,
               "resource_type": "interne", "created_at": now, **payload}
        await db.resources.insert_one(doc)
        return "created"

    if entity == "milestones":
        payload = {k: v for k, v in fields.items() if k != "project_name"}
        if existing:
            await db.milestones.update_one(
                {"milestone_id": existing["milestone_id"]},
                {"$set": {**payload, "updated_at": now}},
            )
            return "updated"
        doc = {
            "milestone_id": str(uuid.uuid4()), "project_id": resolved["project_id"],
            "tenant_id": tid, "status": "planned", "is_governance": False,
            "is_blocking": False, "comment": "", "created_at": now,
            "created_by": user.user_id, **payload,
        }
        await db.milestones.insert_one(doc)
        return "created"

    if entity in ("risks", "decisions"):
        coll = db.risks if entity == "risks" else db.decisions
        id_field = "risk_id" if entity == "risks" else "decision_id"
        payload = {k: v for k, v in fields.items() if k not in ("project_name", "criticality")}
        if entity == "risks":
            prob = payload.get("probability", (existing or {}).get("probability", 3))
            imp = payload.get("impact", (existing or {}).get("impact", 3))
            payload["criticality"] = int(prob) * int(imp)
        if existing:
            await coll.update_one({id_field: existing[id_field]}, {"$set": payload})
            return "updated"
        doc = {
            id_field: str(uuid.uuid4()), "tenant_id": tid,
            "project_id": resolved["project_id"], "created_at": now,
        }
        if entity == "risks":
            doc.update({"category": "autre", "probability": 3, "impact": 3, "status": "identifié"})
        else:
            doc.update({"category": "autre", "status": "proposée"})
        doc.update(payload)
        await coll.insert_one(doc)
        return "created"

    if entity == "demands":
        if existing:
            await db.demands.update_one(
                {"demand_id": existing["demand_id"]}, {"$set": {**fields, "updated_at": now}}
            )
            return "updated"
        doc = {
            "demand_id": str(uuid.uuid4()), "tenant_id": tid, "status": "nouvelle",
            "urgency": "medium", "created_by": user.user_id, "created_by_name": user.name,
            "created_at": now, "updated_at": now, **fields,
        }
        await db.demands.insert_one(doc)
        return "created"

    if entity == "timesheets":
        jh = data.get("jh_value") or 0
        if existing:
            await db.timesheets.update_one(
                {"timesheet_id": existing["timesheet_id"]},
                {"$set": {"jh_value": jh, "status": "draft", "rejection_reason": None,
                          "updated_at": now}},
            )
            return "updated"
        doc = {
            "timesheet_id": str(uuid.uuid4()), "tenant_id": tid,
            "resource_id": resolved["resource_id"],
            "work_allocation_id": resolved["work_allocation_id"],
            "date": data.get("date"), "jh_value": jh, "status": "draft",
            "accounted": False, "submitted_at": None, "cp_reviewed_at": None,
            "validated_at": None, "validated_by": None, "rejection_reason": None,
            "created_at": now,
        }
        await db.timesheets.insert_one(doc)
        return "created"

    raise HTTPException(status_code=400, detail="Entité non gérée")


# ─── Endpoints logiques ──────────────────────────────────────────────────────

def _parse_xlsx(content: bytes, spec: dict) -> list[dict]:
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception:
        raise HTTPException(status_code=422, detail="Fichier illisible : format .xlsx attendu")
    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=422, detail="Fichier vide")

    col_map: dict[int, dict] = {}
    by_norm = {}
    for c in spec["columns"]:
        by_norm[_norm(c["label"])] = c
        by_norm[_norm(c["field"])] = c
    for idx, h in enumerate(header or []):
        c = by_norm.get(_norm(h))
        if c:
            col_map[idx] = c
    if not col_map:
        raise HTTPException(
            status_code=422,
            detail="Aucune colonne reconnue : utilisez le fichier exporté depuis MARCEL comme modèle",
        )

    parsed = []
    for rnum, raw_row in enumerate(rows_iter, start=2):
        if raw_row is None or all(v is None or str(v).strip() == "" for v in raw_row):
            continue
        data: dict = {}
        errors: list[str] = []
        for idx, c in col_map.items():
            raw = raw_row[idx] if idx < len(raw_row) else None
            v, err = _parse_value(raw, c)
            if err:
                errors.append(err)
            elif v is not None:
                data[c["field"]] = v
        parsed.append({"row_num": rnum, "data": data, "parse_errors": errors})
    return parsed


def _normalize_json_rows(rows: list[dict], spec: dict) -> list[dict]:
    parsed = []
    cols = {c["field"]: c for c in spec["columns"]}
    for i, r in enumerate(rows):
        data: dict = {}
        errors: list[str] = []
        for field, raw in (r or {}).items():
            c = cols.get(field)
            if not c:
                continue
            v, err = _parse_value(raw, c)
            if err:
                errors.append(err)
            elif v is not None:
                data[field] = v
        parsed.append({"row_num": i + 1, "data": data, "parse_errors": errors})
    return parsed


async def _analyze(entity: str, parsed: list[dict], user: TokenPayload) -> list[dict]:
    spec = _spec(entity)
    ctx = await _build_ctx(entity, user)
    out = []
    seen_keys: set = set()
    for p in parsed:
        info = _analyze_row(entity, spec, p["data"], ctx)
        errors = p["parse_errors"] + info["errors"]
        mk = tuple(_key(p["data"].get(f)) for f in spec["match"])
        if all(mk) and mk in seen_keys:
            errors.append("Ligne en doublon dans le fichier")
        seen_keys.add(mk)
        out.append({
            "row_num": p["row_num"],
            "action": "error" if errors else info["action"],
            "errors": errors,
            "data": p["data"],
            "_info": info,
        })
    return out


async def preview_import(entity: str, content: bytes, user: TokenPayload) -> dict:
    require_write(user)
    spec = _spec(entity)
    parsed = _parse_xlsx(content, spec)
    analyzed = await _analyze(entity, parsed, user)
    counts = {"new": 0, "update": 0, "error": 0}
    rows = []
    for a in analyzed:
        counts[a["action"]] += 1
        rows.append({k: v for k, v in a.items() if k != "_info"})
    return {
        "entity": entity,
        "label": spec["label"],
        "columns": [{"field": c["field"], "label": c["label"]}
                    for c in spec["columns"] if not c.get("readonly")],
        "rows": rows,
        "counts": counts,
        "total": len(rows),
    }


async def commit_import(entity: str, rows: list[dict], user: TokenPayload) -> dict:
    require_write(user)
    spec = _spec(entity)
    if not rows:
        raise HTTPException(status_code=422, detail="Aucune ligne à importer")
    parsed = _normalize_json_rows(rows, spec)
    analyzed = await _analyze(entity, parsed, user)
    created = updated = skipped = 0
    errors = []
    for a in analyzed:
        if a["action"] == "error":
            skipped += 1
            errors.append({"row": a["row_num"], "messages": a["errors"]})
            continue
        try:
            result = await _apply_row(entity, spec, a["data"], a["_info"], user)
            if result == "created":
                created += 1
            else:
                updated += 1
        except Exception as e:
            skipped += 1
            errors.append({"row": a["row_num"], "messages": [str(e)]})
    return {"entity": entity, "created": created, "updated": updated,
            "skipped": skipped, "errors": errors[:50]}
