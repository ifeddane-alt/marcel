"""Backend tests for Excel export/import & MS Project (.mpp) import (iteration 51)."""
import io
import os
import time
import pytest
import requests
from openpyxl import load_workbook, Workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "http://localhost:8001").rstrip("/")
API = f"{BASE_URL}/api"

CREDS = {"email": "admin@altair.fr", "password": "Admin2026!"}
ENTITIES = ["projects", "programs", "teams", "resources", "milestones",
            "risks", "decisions", "demands", "budget", "timesheets"]


@pytest.fixture(scope="module")
def token():
    r = requests.post(f"{API}/auth/login", json=CREDS, timeout=30)
    assert r.status_code == 200, r.text
    return r.json()["access_token"]


@pytest.fixture(scope="module")
def headers(token):
    return {"Authorization": f"Bearer {token}"}


# ---- Regression: login + projects ----
def test_login_ok(token):
    assert token


def test_projects_list(headers):
    r = requests.get(f"{API}/projects", headers=headers, timeout=30)
    assert r.status_code == 200
    assert isinstance(r.json(), list)


# ---- Excel export for each entity ----
@pytest.mark.parametrize("entity", ENTITIES)
def test_excel_export(entity, headers):
    r = requests.get(f"{API}/excel/{entity}/export", headers=headers, timeout=60)
    assert r.status_code == 200, f"{entity}: {r.status_code} {r.text[:200]}"
    ct = r.headers.get("content-type", "")
    assert "spreadsheetml" in ct or "xlsx" in ct, f"{entity} ct={ct}"
    wb = load_workbook(io.BytesIO(r.content))
    assert len(wb.sheetnames) >= 1
    # Aide import tab
    assert any("Aide" in s or "aide" in s.lower() for s in wb.sheetnames), \
        f"{entity} sheets={wb.sheetnames}"
    ws = wb[wb.sheetnames[0]]
    headers_row = [c.value for c in ws[1] if c.value]
    assert len(headers_row) > 0


# ---- Round trip projects ----
def test_projects_roundtrip(headers):
    # Export
    r = requests.get(f"{API}/excel/projects/export", headers=headers, timeout=60)
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in ws[1]]
    # Find columns
    def col(name):
        for i, h in enumerate(header):
            if h and name.lower() in str(h).lower():
                return i
        return None

    name_col = col("Nom du projet")
    start_col = col("Date début") if col("Date début") is not None else col("Debut")
    end_col = col("Fin prévue") if col("Fin prévue") is not None else col("Fin")
    budget_col = None
    for i, h in enumerate(header):
        if h and "budget" in str(h).lower() and "total" in str(h).lower():
            budget_col = i
            break
    assert name_col is not None and start_col is not None and end_col is not None

    # Count existing rows
    existing_rows = list(ws.iter_rows(min_row=2, values_only=True))
    existing_rows = [r for r in existing_rows if r[name_col]]
    n_existing = len(existing_rows)
    assert n_existing >= 1

    # Modify budget_total of first row (if column present) and remember original
    original_budget = None
    first_project_name = existing_rows[0][name_col]
    if budget_col is not None:
        original_budget = existing_rows[0][budget_col]
        try:
            new_val = (float(original_budget) if original_budget else 0) + 1234
        except Exception:
            new_val = 12345
        ws.cell(row=2, column=budget_col + 1, value=new_val)

    # Add new project row
    new_row_idx = ws.max_row + 1
    ws.cell(row=new_row_idx, column=name_col + 1, value="TEST ROUNDTRIP PROJ")
    ws.cell(row=new_row_idx, column=start_col + 1, value="01/01/2026")
    ws.cell(row=new_row_idx, column=end_col + 1, value="31/12/2026")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # Preview
    files = {"file": ("test.xlsx", buf.getvalue(),
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    rp = requests.post(f"{API}/excel/projects/import/preview",
                       headers=headers, files=files, timeout=60)
    assert rp.status_code == 200, rp.text
    preview = rp.json()
    counts = preview.get("counts", {})
    assert counts.get("new", 0) >= 1, f"preview: {preview}"
    assert counts.get("update", 0) >= 1
    assert counts.get("error", 0) == 0, f"errors: {preview}"

    # Commit valid rows only (extract data)
    rows = [r["data"] for r in preview.get("rows", []) if r.get("action") != "error"]
    rc = requests.post(f"{API}/excel/projects/import/commit",
                       headers=headers, json={"rows": rows}, timeout=90)
    assert rc.status_code == 200, rc.text
    commit_result = rc.json()
    assert commit_result.get("created", 0) >= 1, f"commit: {commit_result}"

    # Verify new project exists
    rl = requests.get(f"{API}/projects", headers=headers, timeout=30)
    assert rl.status_code == 200
    projects = rl.json()
    test_proj = next((p for p in projects if p.get("name") == "TEST ROUNDTRIP PROJ"), None)
    assert test_proj is not None, "New project not created"

    # Cleanup: delete new project
    del_r = requests.delete(f"{API}/projects/{test_proj['project_id']}", headers=headers, timeout=30)
    assert del_r.status_code in (200, 204)

    # Restore original budget if we changed one
    if budget_col is not None and original_budget is not None:
        first = next((p for p in projects if p.get("name") == first_project_name), None)
        if first:
            try:
                requests.patch(f"{API}/projects/{first['project_id']}", headers=headers,
                               json={"budget_total": float(original_budget)}, timeout=30)
            except Exception:
                pass


# ---- Risks import with error ----
def test_risks_import_error(headers):
    r = requests.get(f"{API}/excel/risks/export", headers=headers, timeout=60)
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in ws[1]]

    proj_col = None
    title_col = None
    for i, h in enumerate(header):
        if not h:
            continue
        h_low = str(h).lower()
        if "projet" in h_low and proj_col is None:
            proj_col = i
        if ("titre" in h_low or "libell" in h_low or "risque" in h_low) and title_col is None:
            title_col = i
    assert proj_col is not None

    new_row = ws.max_row + 1
    ws.cell(row=new_row, column=proj_col + 1, value="PROJET_INEXISTANT_ZZZ_XYZ")
    if title_col is not None:
        ws.cell(row=new_row, column=title_col + 1, value="Risque test invalide")

    buf = io.BytesIO()
    wb.save(buf)
    files = {"file": ("r.xlsx", buf.getvalue(),
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    rp = requests.post(f"{API}/excel/risks/import/preview",
                       headers=headers, files=files, timeout=60)
    assert rp.status_code == 200, rp.text
    preview = rp.json()
    # Find the erroneous row
    error_rows = [row for row in preview.get("rows", []) if row.get("action") == "error"]
    assert len(error_rows) >= 1, f"Expected error row, got: {preview}"
    joined_msgs = " ".join([" ".join(r.get("errors", [])) for r in error_rows]).lower()
    assert "projet" in joined_msgs and ("introuv" in joined_msgs or "inconnu" in joined_msgs), \
        f"messages: {joined_msgs}"


# ---- Budget update-only ----
def test_budget_update_only(headers):
    # Export budget
    r = requests.get(f"{API}/excel/budget/export", headers=headers, timeout=60)
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in ws[1]]

    def find(*keys):
        for i, h in enumerate(header):
            if not h:
                continue
            hl = str(h).lower()
            if all(k in hl for k in keys):
                return i
        return None

    proj_col = find("projet")
    capex_col = find("capex")
    opex_col = find("opex")
    assert proj_col is not None

    # Add row for unknown project => must be error
    new_row = ws.max_row + 1
    ws.cell(row=new_row, column=proj_col + 1, value="XXXX_UNKNOWN_BUDGET_PROJ")
    if capex_col is not None:
        ws.cell(row=new_row, column=capex_col + 1, value=1000)

    # Modify an existing project (first row) if capex/opex present, remember originals
    original_row1 = {}
    p1_name = ws.cell(row=2, column=proj_col + 1).value
    if capex_col is not None:
        original_row1["capex"] = ws.cell(row=2, column=capex_col + 1).value
        try:
            new_v = (float(original_row1["capex"]) if original_row1["capex"] else 0) + 500
        except Exception:
            new_v = 5000
        ws.cell(row=2, column=capex_col + 1, value=new_v)
    if opex_col is not None:
        original_row1["opex"] = ws.cell(row=2, column=opex_col + 1).value
        try:
            new_v = (float(original_row1["opex"]) if original_row1["opex"] else 0) + 300
        except Exception:
            new_v = 3000
        ws.cell(row=2, column=opex_col + 1, value=new_v)

    buf = io.BytesIO()
    wb.save(buf)
    files = {"file": ("b.xlsx", buf.getvalue(),
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    rp = requests.post(f"{API}/excel/budget/import/preview",
                       headers=headers, files=files, timeout=60)
    assert rp.status_code == 200, rp.text
    preview = rp.json()
    # Must have at least 1 error (unknown project) and 0 new (update-only)
    counts = preview.get("counts", {})
    assert counts.get("new", 0) == 0, f"budget should not create: {preview}"
    assert counts.get("error", 0) >= 1
    error_msgs = " ".join([" ".join(r.get("errors", [])) for r in preview.get("rows", [])
                           if r.get("action") == "error"]).lower()
    assert "budget" in error_msgs or "cr" in error_msgs  # message about "ne crée pas"

    # Commit valid updates
    rows = [r["data"] for r in preview.get("rows", []) if r.get("action") == "update"]
    if rows:
        rc = requests.post(f"{API}/excel/budget/import/commit",
                           headers=headers, json={"rows": rows}, timeout=60)
        assert rc.status_code == 200

    # Restore original values if we changed
    if p1_name and (capex_col is not None or opex_col is not None):
        pl = requests.get(f"{API}/projects", headers=headers, timeout=30).json()
        p = next((x for x in pl if x.get("name") == p1_name), None)
        if p:
            patch = {}
            if "capex" in original_row1 and original_row1["capex"] is not None:
                try:
                    patch["capex_planned"] = float(original_row1["capex"])
                except Exception:
                    pass
            if "opex" in original_row1 and original_row1["opex"] is not None:
                try:
                    patch["opex_planned"] = float(original_row1["opex"])
                except Exception:
                    pass
            if patch:
                requests.patch(f"{API}/projects/{p['project_id']}", headers=headers, json=patch, timeout=30)


# ---- MS Project .mpp import ----
def test_msproject_mpp_import(headers):
    # Create temp project
    payload = {
        "name": "TEST MPP IMPORT",
        "start_date": "2026-01-01",
        "end_date": "2026-12-31",
        "methodology": "waterfall",
        "status_rag": "green",
        "jh_planned": 100,
        "end_date_baseline": "2026-12-31",
        "end_date_forecast": "2026-12-31",
    }
    cr = requests.post(f"{API}/projects", headers=headers, json=payload, timeout=30)
    assert cr.status_code in (200, 201), cr.text
    proj = cr.json()
    pid = proj["project_id"]
    try:
        assert os.path.exists("/tmp/sample.mpp")
        with open("/tmp/sample.mpp", "rb") as f:
            content = f.read()
        # Verify OLE signature
        assert content[:4] == b"\xd0\xcf\x11\xe0"
        files = {"file": ("sample.mpp", content, "application/octet-stream")}
        t0 = time.time()
        r = requests.post(f"{API}/msproject/import/{pid}", headers=headers,
                          files=files, timeout=120)
        elapsed = time.time() - t0
        assert r.status_code == 200, f"status={r.status_code} body={r.text[:500]}"
        data = r.json()
        assert data.get("tasks_created", 0) > 0, f"resp={data}"
        print(f"MPP import: {data.get('tasks_created')} tasks in {elapsed:.1f}s")
    finally:
        requests.delete(f"{API}/projects/{pid}", headers=headers, timeout=30)
